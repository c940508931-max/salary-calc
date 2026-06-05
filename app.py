"""
薪资计算工具 - Flask 本地应用
底薪+绩效+提成+补贴，按自然日折算，自动算社保
"""
import calendar
import json
import logging
import os
import time
import uuid
from datetime import datetime, date
from io import BytesIO
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    send_file, flash, session as flask_session
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", BASE_DIR / "uploads"))
TEMPLATE_DIR = BASE_DIR / "templates"
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "salary-calc-secret-key-change-in-production")
app.config["UPLOAD_FOLDER"] = str(UPLOAD_DIR)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

# 预览数据缓存（文件存储，避免 session 大小限制）
PREVIEW_CACHE_DIR = BASE_DIR / "_preview_cache"
os.makedirs(PREVIEW_CACHE_DIR, exist_ok=True)
CACHE_MAX_AGE = 3600  # 1小时后清理


def _cleanup_old_files():
    """清理超过1小时的旧上传文件和缓存"""
    now = time.time()
    for d, max_age in [(UPLOAD_DIR, CACHE_MAX_AGE), (PREVIEW_CACHE_DIR, CACHE_MAX_AGE)]:
        for f in d.iterdir():
            if f.is_file() and now - f.stat().st_mtime > max_age:
                try:
                    f.unlink()
                except OSError:
                    pass

# 社保比例（个人部分）
SOCIAL_INSURANCE_RATES = {
    "养老保险": 0.08,   # 8%
    "医疗保险": 0.02,   # 2%
    "失业保险": 0.005,  # 0.5%
}

# 补贴选项
SUBSIDY_OPTIONS = ["餐补", "交通补贴", "住房补贴", "通讯补贴"]

# 常用样式常量
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 模板列定义（对应导出时的表头）
# (列名, 宽度)
TEMPLATE_COLUMNS = [
    ("姓名", 10),
    ("工号", 10),
    ("身份证号", 20),
    ("入职日期", 14),
    ("转正日期", 14),
    ("手机号", 16),
    ("应出勤天数", 12),
    ("请假天数", 12),
    ("调休天数", 12),
    ("实际出勤天数", 14),
    ("基本工资", 12),
    ("转正薪资", 12),
    ("试用期薪资", 12),
    ("岗位绩效", 12),
    ("加班费", 10),
    ("销售提成", 12),
    ("补卡次数", 10),
    ("迟到早退(分钟)", 14),
]

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def get_month_days(dt: date) -> int:
    """获取当月总自然天数"""
    return calendar.monthrange(dt.year, dt.month)[1]


def get_work_days_from_hire(hire_date: date, ref_date: date) -> int:
    """
    计算应出勤天数：
    - 尚未入职：返回 0
    - 入职当月：从入职日到当月底（含入职日）
    - 入职后月份：整月全勤
    比如 4月24入职，算4月=7天，算5月=31天
    """
    ref_last_day = calendar.monthrange(ref_date.year, ref_date.month)[1]
    ref_month_end = date(ref_date.year, ref_date.month, ref_last_day)

    # 入职日在参考月份之后 → 尚未入职
    if hire_date > ref_month_end:
        return 0

    # 入职日期 ≥ 参考月份首日 → 入职当月
    ref_month_start = ref_date.replace(day=1)
    if hire_date >= ref_month_start:
        last_day = calendar.monthrange(hire_date.year, hire_date.month)[1]
        month_end = date(hire_date.year, hire_date.month, last_day)
        return (month_end - hire_date).days + 1

    # 入职日在参考月份之前 → 整月全勤
    return ref_last_day


def calc_proportional_salary(monthly_salary: float, work_days: int, month_days: int) -> float:
    """按自然日折算工资：月薪 ÷ 当月总天数 × 实际出勤天数"""
    if month_days == 0:
        return 0
    return round(monthly_salary / month_days * work_days, 2)


def calc_social_insurance(base: float) -> dict:
    """计算社保各项"""
    result = {}
    for name, rate in SOCIAL_INSURANCE_RATES.items():
        result[name] = round(base * rate, 2)
    return result


def parse_date(value) -> date | None:
    """尝试解析各种日期格式"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_float(value) -> float:
    """安全转浮点"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip().replace(",", "").replace("￥", "").replace("¥", "")
        if value in ("", "/", "-", "--"):
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0
    return 0.0


def _style_header(ws, row, col_count):
    """给表头行设置样式"""
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _THIN_BORDER


def _style_data_cell(cell, is_date=False):
    """给数据单元格设置样式"""
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER
    if is_date:
        cell.number_format = "YYYY-MM-DD"


def create_template(workbook=None) -> Workbook:
    """创建导入模板 Excel，带填写说明行"""
    wb = workbook or Workbook()
    ws = wb.active
    ws.title = "工资表"

    # 构建表头（第1行）
    headers = list(TEMPLATE_COLUMNS)  # [(name, width), ...]
    # 插入补贴列
    full_headers = []
    for i, (name, width) in enumerate(headers):
        full_headers.append((name, width))
        if name == "销售提成":
            for sub in SUBSIDY_OPTIONS:
                full_headers.append((sub, 12))

    # 追加后面的固定列
    fixed_tail = [
        ("社保基数", 12),
        ("养老保险", 12),
        ("医疗保险", 12),
        ("失业保险", 12),
        ("住房公积金", 12),
        ("个人所得税", 12),
        ("全勤奖", 10),
        ("补卡扣款", 10),
        ("迟到扣款", 10),
        ("实发金额", 12),
    ]
    full_headers.extend(fixed_tail)

    # 确认表头数与类型数一致
    assert len(full_headers) == 32, f"表头数应为32，实际{len(full_headers)}"

    # 写标题行（第1行）
    ws.cell(row=1, column=1, value="路易小姐薪资表 - 📌 浅蓝底色 = 自动计算（不要填写）| 白色底色 = 请手动填写")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(full_headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- 填写说明（第2行）：标注每列是 手动填 还是 自动算 ----
    # 定义各列说明及颜色
    # "M" = 手动填写, "A" = 自动计算
    col_types = [
        "M",       # 姓名
        "M",       # 工号
        "M",       # 身份证号
        "M",       # 入职日期
        "M",       # 转正日期（可留空）
        "M",       # 手机号
        "A",       # 应出勤天数（自动）
        "M",       # 请假天数
        "M",       # 调休天数
        "A",       # 实际出勤天数（自动：应出勤+调休-请假）
        "A",       # 基本工资（自动）
        "M",       # 转正薪资
        "M",       # 试用期薪资（可留空）
        "M",       # 岗位绩效
        "M",       # 加班费
        "M",       # 销售提成
        "M",       # 补卡次数
        "M",       # 迟到早退(分钟)
    ]
    # 补贴列（手动填）
    for _ in SUBSIDY_OPTIONS:
        col_types.append("M")
    # 尾部
    tail_types = [
        "A",       # 社保基数（跟随网页设置）
        "A",       # 养老保险
        "A",       # 医疗保险
        "A",       # 失业保险
        "M",       # 住房公积金
        "M",       # 个人所得税
        "A",       # 全勤奖（自动）
        "A",       # 补卡扣款（自动）
        "A",       # 迟到扣款（自动）
        "A",       # 实发金额（自动）
    ]
    col_types.extend(tail_types)

    # 说明文字
    col_hints = {
        "M": "✏️ 手动填写",
        "A": "⚡ 自动计算",
    }

    # 填色
    manual_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")       # 白色
    auto_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")        # 浅灰蓝

    # 第2行：列名
    # 第3行：说明行
    for col_idx, ((name, width), ctype) in enumerate(zip(full_headers, col_types), 1):
        # 列名（第2行）— 手动列白底深字，自动列浅灰蓝底白字
        cell = ws.cell(row=2, column=col_idx, value=name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if ctype == "A":
            cell.fill = auto_fill
            cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        else:
            cell.fill = manual_fill
            cell.font = Font(name="微软雅黑", bold=True, size=10, color="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    # 说明行（第3行）
    hint_font = Font(name="微软雅黑", bold=True, size=9, italic=True)
    manual_text_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 浅绿
    auto_text_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")    # 浅粉
    for col_idx, ctype in enumerate(col_types, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_hints[ctype])
        cell.font = hint_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = auto_text_fill if ctype == "A" else manual_text_fill
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 22

    # 图例说明行（第4行）
    legend_cell = ws.cell(row=4, column=1,
        value="🟢 浅绿 = 请手动填写  |  🩷 浅粉 = 系统自动计算，请勿填写")
    legend_cell.font = Font(name="微软雅黑", size=9, color="666666")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(full_headers))
    ws.row_dimensions[4].height = 20

    # 示例数据行（第5行）
    example = [
        "张三", "LY0001", "330102199001011234",
        date(2026, 4, 1), "",          # 入职日期, 转正日期（空=默认无试用期）
        "+86 13800138000",
        None, 0, 0, None, None, 15000, # 应出勤=自动, 请假=0, 调休=0, 实际出勤=自动, 基本工资=自动, 转正薪资=15000
        "",                            # 试用期薪资（空=无试用期）
        300, 200, 0,                   # 岗位绩效, 加班费, 销售提成
        0, 0,                          # 补卡次数, 迟到早退(分钟)
    ]
    subsidy_example = [300, 200, 0, 0]
    tail_example = [5000, None, None, None, 0, 240, None, None, None, None]

    row_data = example + subsidy_example + tail_example
    for col_idx, (val, ctype) in enumerate(zip(row_data, col_types), 1):
        if ctype == "A":
            # 自动计算的列显示灰色文字说明
            cell = ws.cell(row=5, column=col_idx,
                value="=(自动)" if val is None else val)
            cell.font = Font(name="微软雅黑", size=10, color="999999", italic=True)
            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        else:
            cell = ws.cell(row=5, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        if col_idx == 4:
            cell.number_format = "YYYY-MM-DD"
    ws.row_dimensions[5].height = 22

    # 冻结窗格（冻结前4行）
    ws.freeze_panes = "A5"

    return wb


# ---------------------------------------------------------------------------
# 解析导入的数据
# ---------------------------------------------------------------------------

def parse_uploaded_excel(filepath: str, social_base: float, ref_month: str) -> list[dict]:
    """
    解析上传的 Excel，返回工资数据列表
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # 读取表头（第2行）
    headers = []
    for cell in ws[2]:
        headers.append(str(cell.value or "").strip())

    # 找到各列的索引
    col_map = {}
    for idx, h in enumerate(headers):
        h_lower = h.lower().replace(" ", "")
        if h_lower in ("姓名", "工号", "身份证号", "入职日期", "转正日期", "手机号",
                        "应出勤天数", "请假/调休天数", "请假天数", "调休天数", "实际出勤天数",
                        "基本工资", "总薪资", "转正薪资", "试用期薪资", "岗位绩效",
                        "加班费", "销售提成", "补卡次数", "迟到早退(分钟)",
                        "社保基数", "养老保险", "医疗保险",
                        "失业保险", "住房公积金", "个人所得税", "实发金额",
                        "全勤奖", "补卡扣款", "迟到扣款"):
            col_map[h] = idx
        elif h in SUBSIDY_OPTIONS:
            col_map.setdefault("补贴列", []).append(idx)

    subsidy_indices = col_map.get("补贴列", [])

    # 解析数据行（从第3行开始）
    results = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = str(row[col_map.get("姓名", 0)] or "").strip() if col_map.get("姓名", 0) < len(row) else ""
        if not name:
            continue

        hire_date = parse_date(row[col_map.get("入职日期", 3)] if col_map.get("入职日期", 3) < len(row) else None)

        if hire_date is None:
            continue

        # 确定参考月份
        if ref_month:
            try:
                ref = datetime.strptime(ref_month, "%Y-%m").date()
            except ValueError:
                ref = hire_date
        else:
            ref = hire_date

        month_days = get_month_days(ref)

        # 应出勤天数 = 根据入职日和参考月份计算
        scheduled_days = get_work_days_from_hire(hire_date, ref)

        # 请假天数（手动填，兼容旧模板的"请假/调休天数"）
        leave_idx = col_map.get("请假天数")
        if leave_idx is None:
            leave_idx = col_map.get("请假/调休天数", 7)
        leave_days = int(parse_float(row[leave_idx] if leave_idx is not None and leave_idx < len(row) else 0))

        # 调休天数（手动填）
        comp_idx = col_map.get("调休天数", 99)
        comp_days = int(parse_float(row[comp_idx] if comp_idx < len(row) else 0))

        # 实际出勤天数 = 应出勤 + 调休 - 请假（自动，最小0）
        actual_days = max(scheduled_days + comp_days - leave_days, 0)

        # 转正薪资（手动填，兼容旧模板的"总薪资"）
        total_salary_idx = col_map.get("转正薪资")
        if total_salary_idx is None:
            total_salary_idx = col_map.get("总薪资", 11)
        total_salary = parse_float(row[total_salary_idx] if total_salary_idx is not None and total_salary_idx < len(row) else 0)

        # ===== 试用期 / 转正判断 =====
        probation_salary = parse_float(row[col_map.get("试用期薪资", 99)] if col_map.get("试用期薪资", 99) < len(row) else 0)
        regular_idx = col_map.get("转正日期", 99)
        regular_date = parse_date(row[regular_idx] if regular_idx < len(row) else None)

        has_probation = probation_salary > 0 and regular_date is not None
        effective_salary = total_salary
        probation_note = ""

        if has_probation:
            ref_month_start = ref.replace(day=1)
            _, last_day = calendar.monthrange(ref.year, ref.month)
            ref_month_end = ref.replace(day=last_day)

            if regular_date > ref_month_end:
                effective_salary = probation_salary
                probation_note = "试用期"
            elif regular_date >= ref_month_start:
                probation_days = (regular_date - ref_month_start).days
                regular_days = month_days - probation_days
                effective_salary = (probation_salary * probation_days + total_salary * regular_days) / month_days
                probation_note = f"试用{probation_days}天+转正{regular_days}天"

        # ===== 全勤奖拆分 =====
        # 转正薪资拆分出 200 全勤奖
        base_salary = max(0, effective_salary - 200)

        # 基本工资 = 底薪 ÷ 当月总天数 × 实际出勤天数（自动）
        base_pay = calc_proportional_salary(base_salary, actual_days, month_days)

        # ===== 补卡次数 =====
        card_idx = col_map.get("补卡次数", 99)
        comp_card_count = int(parse_float(row[card_idx] if card_idx < len(row) else 0))

        # ===== 迟到早退 =====
        late_idx = col_map.get("迟到早退(分钟)", 99)
        actual_late_minutes = parse_float(row[late_idx] if late_idx < len(row) else 0)

        # 绩效
        performance = parse_float(row[col_map.get("岗位绩效", 13)] if col_map.get("岗位绩效", 13) < len(row) else 0)

        # 加班费
        overtime = parse_float(row[col_map.get("加班费", 14)] if col_map.get("加班费", 14) < len(row) else 0)

        # 销售提成
        commission = parse_float(row[col_map.get("销售提成", 15)] if col_map.get("销售提成", 15) < len(row) else 0)

        # 补贴（动态列）
        subsidies = {}
        for idx in subsidy_indices:
            if idx < len(row):
                sub_name = headers[idx]
                sub_val = parse_float(row[idx])
                if sub_val > 0:
                    subsidies[sub_name] = sub_val

        subsidy_total = sum(subsidies.values())

        # ==== 全勤奖 ====
        # 请假 > 0 或 补卡 > 2 → 全勤奖 = 0
        full_attendance_bonus = 200
        if leave_days > 0 or comp_card_count > 2:
            full_attendance_bonus = 0

        # ==== 补卡扣款 ====
        # 超 2 次后每超 1 次扣半天底薪
        daily_base_salary = round(base_salary / month_days, 4) if month_days > 0 else 0
        comp_card_deduction = 0
        if comp_card_count > 2:
            extra_cards = comp_card_count - 2
            comp_card_deduction = round(extra_cards * (daily_base_salary * 0.5), 2)

        # ==== 迟到早退扣款 ====
        # 入职15号(含)内允许150分钟，15号后允许75分钟
        if hire_date.day <= 15:
            allowed_late = 150
        else:
            allowed_late = 75
        late_excess = max(0, actual_late_minutes - allowed_late)
        late_deduction = round(late_excess * 2, 2)

        # 社保 —
        if (hire_date.year == ref.year and hire_date.month == ref.month):
            if hire_date.day <= 15:
                social = calc_social_insurance(social_base)
            else:
                social = {"养老保险": 0, "医疗保险": 0, "失业保险": 0}
        else:
            social = calc_social_insurance(social_base)

        social_total = social["养老保险"] + social["医疗保险"] + social["失业保险"]

        housing_fund = parse_float(row[col_map.get("住房公积金", 20)] if col_map.get("住房公积金", 20) < len(row) else 0)
        tax = parse_float(row[col_map.get("个人所得税", 21)] if col_map.get("个人所得税", 21) < len(row) else 0)

        # 应发合计 = 基本工资 + 全勤奖 + 绩效 + 加班费 + 提成 + 补贴
        gross = round(base_pay + full_attendance_bonus + performance + overtime + commission + subsidy_total, 2)
        # 扣款合计 = 社保 + 公积金 + 个税 + 补卡扣款 + 迟到扣款
        deductions = round(social_total + housing_fund + tax + comp_card_deduction + late_deduction, 2)
        # 实发
        net = round(gross - deductions, 2)

        employee_id = str(row[col_map.get("工号", 1)] or "").strip() if col_map.get("工号", 1) < len(row) else ""
        id_card = str(row[col_map.get("身份证号", 2)] or "").strip() if col_map.get("身份证号", 2) < len(row) else ""
        phone = str(row[col_map.get("手机号", 4)] or "").strip() if col_map.get("手机号", 4) < len(row) else ""

        results.append({
            "name": name,
            "employee_id": employee_id,
            "id_card": id_card,
            "hire_date": hire_date,
            "regular_date": regular_date,
            "probation_salary": probation_salary,
            "effective_salary": round(effective_salary, 2),
            "probation_note": probation_note,
            "phone": phone,
            "scheduled_days": scheduled_days,
            "leave_days": leave_days,
            "comp_days": comp_days,
            "actual_days": actual_days,
            "month_days": month_days,
            "base_salary": round(base_salary, 2),
            "base_pay": base_pay,
            "total_salary": total_salary,
            "full_attendance_bonus": full_attendance_bonus,
            "comp_card_count": comp_card_count,
            "comp_card_deduction": comp_card_deduction,
            "actual_late_minutes": actual_late_minutes,
            "allowed_late": allowed_late,
            "late_excess": late_excess,
            "late_deduction": late_deduction,
            "performance": performance,
            "overtime": overtime,
            "commission": commission,
            "subsidies": subsidies,
            "subsidy_total": subsidy_total,
            "social_base": social_base,
            "social_pension": social["养老保险"],
            "social_medical": social["医疗保险"],
            "social_unemployment": social["失业保险"],
            "housing_fund": housing_fund,
            "tax": tax,
            "gross": gross,
            "deductions": deductions,
            "net": net,
            "ref_month": ref,
        })

    wb.close()
    return results


# ---------------------------------------------------------------------------
# 导出计算结果到 Excel
# ---------------------------------------------------------------------------

def export_to_excel(data: list[dict], social_base: float) -> BytesIO:
    """将计算结果导出为 Excel（与你模板样式一致）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资表"

    # 收集所有补贴列名
    all_subsidies = []
    for d in data:
        for k in d.get("subsidies", {}):
            if k not in all_subsidies:
                all_subsidies.append(k)

    # 构建完整表头
    base_headers = [h[0] for h in TEMPLATE_COLUMNS]
    # 在销售提成之后插入补贴列
    result_headers = []
    for h in base_headers:
        result_headers.append(h)
        if h == "销售提成":
            result_headers.extend(all_subsidies)

    tail_headers = ["社保基数", "养老保险", "医疗保险", "失业保险", "住房公积金", "个人所得税",
                     "全勤奖", "补卡扣款", "迟到扣款", "实发金额"]
    result_headers.extend(tail_headers)

    # 写标题行
    ws.cell(row=1, column=1, value="路易小姐薪资表")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(result_headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 写表头行
    for col_idx, name in enumerate(result_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    _style_header(ws, 2, len(result_headers))

    # 写数据行
    for row_idx, d in enumerate(data, 3):
        row_data = [
            d["name"],
            d["employee_id"],
            d["id_card"],
            d["hire_date"],
            d.get("regular_date") or "/",
            d["phone"],
            d["scheduled_days"],
            d.get("leave_days", 0),
            d.get("comp_days", 0),
            d["actual_days"],
            d["base_pay"],
            d.get("total_salary", 0),
            d.get("probation_salary", 0),
            d["performance"],
            d["overtime"],
            d["commission"],
            d.get("comp_card_count", 0),
            d.get("actual_late_minutes", 0),
        ]
        # 补贴
        for sub in all_subsidies:
            row_data.append(d.get("subsidies", {}).get(sub, 0))
        # 尾部
        row_data.extend([
            social_base,
            d["social_pension"],
            d["social_medical"],
            d["social_unemployment"],
            d["housing_fund"] if d["housing_fund"] > 0 else "/",
            d["tax"],
            d.get("full_attendance_bonus", 0),
            d.get("comp_card_deduction", 0),
            d.get("late_deduction", 0),
            d["net"],
        ])

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            is_date = col_idx == 4
            _style_data_cell(cell, is_date=is_date)

    # 冻结窗格
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output


# ---------------------------------------------------------------------------
# Flask 路由
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    """首页"""
    return render_template("index.html",
                         subsidy_options=SUBSIDY_OPTIONS,
                         social_insurance_rates=SOCIAL_INSURANCE_RATES)


@app.route("/download_template")
def download_template():
    """下载 Excel 模板"""
    wb = create_template()
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="薪资模板.xlsx",
    )


@app.route("/preview", methods=["GET", "POST"])
def preview():
    if request.method == "GET":
        return redirect(url_for("index"))
    """上传 Excel 并预览计算结果"""
    social_base = parse_float(request.form.get("social_base", 5000))
    ref_month = request.form.get("ref_month", "")

    if "file" not in request.files:
        flash("请上传文件", "error")
        return redirect(url_for("index"))

    file = request.files["file"]
    if file.filename == "":
        flash("请选择文件", "error")
        return redirect(url_for("index"))

    if not file.filename.endswith((".xlsx", ".xls")):
        flash("请上传 .xlsx 或 .xls 文件", "error")
        return redirect(url_for("index"))

    # 清理旧文件
    _cleanup_old_files()

    # 保存上传文件
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    save_path = UPLOAD_DIR / f"{timestamp}_{file.filename}"
    file.save(str(save_path))

    try:
        data = parse_uploaded_excel(str(save_path), social_base, ref_month)
    except Exception as e:
        logger.exception("解析上传文件失败")
        flash(f"解析文件出错：{str(e)}", "error")
        return redirect(url_for("index"))

    if not data:
        flash("未找到有效数据，请检查模板格式", "error")
        return redirect(url_for("index"))

    # 提取所有出现的补贴列名
    all_subsidies = []
    for d in data:
        for k in d.get("subsidies", {}):
            if k not in all_subsidies:
                all_subsidies.append(k)

    # 缓存到文件（避免 session 大小限制）
    cache_key = str(uuid.uuid4())
    cache_data = {
        "data": [
            {k: (v.isoformat() if isinstance(v, date) else v) for k, v in d.items()}
            for d in data
        ],
        "social_base": social_base,
        "ref_month": ref_month,
    }
    cache_path = PREVIEW_CACHE_DIR / f"{cache_key}.json"
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False)
    flask_session["preview_cache_key"] = cache_key

    return render_template("preview.html",
                         data=data,
                         all_subsidies=all_subsidies,
                         social_base=social_base,
                         social_insurance_rates=SOCIAL_INSURANCE_RATES,
                         ref_month=ref_month,
                         cache_key=cache_key)


@app.route("/export")
def export():
    """导出计算结果"""
    cache_key = request.args.get("key") or flask_session.get("preview_cache_key")
    if not cache_key:
        flash("没有可导出的数据", "error")
        return redirect(url_for("index"))

    cache_path = PREVIEW_CACHE_DIR / f"{cache_key}.json"
    if not cache_path.exists():
        flash("数据已过期，请重新上传", "error")
        return redirect(url_for("index"))

    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cache_data = json.load(f)
    except Exception as e:
        logger.exception("读取缓存数据失败")
        flash("数据读取失败，请重新上传", "error")
        return redirect(url_for("index"))

    raw = cache_data.get("data", [])
    social_base = cache_data.get("social_base", 5000)
    ref_month = cache_data.get("ref_month", "")

    if not raw:
        flash("没有可导出的数据", "error")
        return redirect(url_for("index"))

    # 恢复 date 对象
    data = []
    for d in raw:
        if isinstance(d.get("hire_date"), str):
            d["hire_date"] = date.fromisoformat(d["hire_date"])
        if isinstance(d.get("regular_date"), str) and d.get("regular_date"):
            d["regular_date"] = date.fromisoformat(d["regular_date"])
        if isinstance(d.get("ref_month"), str):
            d["ref_month"] = date.fromisoformat(d["ref_month"])
        data.append(d)

    output = export_to_excel(data, social_base)

    # 生成含月份的文件名
    filename = "薪资计算结果.xlsx"
    if ref_month:
        try:
            ref_dt = datetime.strptime(ref_month, "%Y-%m")
            filename = f"薪资计算结果_{ref_dt.year}年{ref_dt.month}月.xlsx"
        except ValueError:
            pass

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"🚀 薪资计算工具已启动：http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
